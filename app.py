# -*- coding: utf-8 -*-
import os
import io
import streamlit as st
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

st.set_page_config(page_title="Générateur Carte Habilitation ONCF", layout="centered")

def load_excel_data(file_path):
    if not os.path.exists(file_path):
        return {"engins": "", "sites": "", "manoeuvre": "", "titre_engins": "Autorisé à conduire les locos et rames suivantes"}
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    engins, sites, manoeuvre = "", "", ""
    titre_engins = "Autorisé à conduire les locos et rames suivantes"
    
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or "").strip()
            if "E1450" in val or "DH" in val or "Z2M" in val:
                engins = val
            elif "Site" in val or "Lignes" in val or "Réseau" in val:
                sites = val
            elif "Matériel" in val:
                manoeuvre = val
            if "arrêter" in val.lower():
                titre_engins = "Autorisé à arrêter les locos et rames suivantes"
                
    return {
        "engins": engins, 
        "sites": sites, 
        "manoeuvre": manoeuvre,
        "titre_engins": titre_engins
    }

FUNCTIONS_CONFIG = {
    "Chef Formation Trains": "data/Cartes d'habilitation CFT.xlsx",
    "Chef de Train": "data/carte d'habilitation CTR.xlsx",
    "Conducteur de Manœuvre": "data/carte d'habilitation CRMV.xlsx",
    "Conducteur de Ligne": "data/carte d'habilitation CL.xlsx"
}

def register_font():
    candidates = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibri.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("MyFont", p))
                return "MyFont"
            except Exception:
                pass
    return "Helvetica"

FONT = register_font()

def wrap_text(text, max_chars):
    if not text:
        return [""]
    words = str(text).split()
    lines, current = [], ""
    for word in words:
        test = word if not current else current + " " + word
        if len(test) <= max_chars:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines or [""]

def generate_pdf_bytes(data, photo_bytes, excel_info):
    buffer = io.BytesIO()
    CARD_W, CARD_H = 1000, 500
    c = canvas.Canvas(buffer, pagesize=(CARD_W, CARD_H))

    # 1. إطار الكارطة الخارجي
    c.setLineWidth(3)
    c.rect(10, 10, CARD_W - 20, CARD_H - 20)

    # 2. تقسيم الخطوط العمودية والأفقية
    c.setLineWidth(1.5)
    c.line(500, 10, 500, CARD_H - 10) # خط عمودي وسطاني
    c.line(750, 120, 750, CARD_H - 10) # خط عمودي بين Engins و Sites

    c.line(500, 120, CARD_W - 10, 120) # خط تنبيهات السفلي
    if data['manoeuvre']:
        c.line(500, 270, 750, 270) # خط الفاصل ديال Manœuvre

    # 3. الهيدر (ONCF & Titre)
    c.setFont(FONT, 18)
    c.setFillColorRGB(0.85, 0.35, 0) # لون برتقالي
    c.drawString(40, CARD_H - 45, "ONCF")
    
    c.setFont(FONT, 10)
    c.setFillColorRGB(0, 0, 0)
    c.drawString(30, CARD_H - 60, "PV/DTV/EPTCN")

    c.setFont(FONT, 20)
    c.drawString(200, CARD_H - 45, "Titre d'habilitation")
    
    c.setFont(FONT, 16)
    c.setFillColorRGB(0.85, 0.35, 0)
    c.drawString(200, CARD_H - 70, data['fonction'])

    # 4. إطار صورة الشخص
    px, py, pw, ph = 30, CARD_H - 250, 130, 160
    c.rect(px, py, pw, ph)
    if photo_bytes:
        try:
            c.drawImage(ImageReader(io.BytesIO(photo_bytes)), px + 2, py + 2, width=pw - 4, height=ph - 4, preserveAspectRatio=True)
        except Exception:
            pass

    # 5. المعلومات الشخصية
    c.setFillColorRGB(0, 0, 0)
    c.setFont(FONT, 12)
    
    labels_y = [
        ("Nom :", data['nom'], CARD_H - 120),
        ("Prénom :", data['prenom'], CARD_H - 145),
        ("Matricule :", data['matricule'], CARD_H - 170),
        ("Centre :", data['centre'], CARD_H - 195),
        ("Antenne :", data['antenne'], CARD_H - 220),
        ("Date d'autorisation :", data['date_aut'], CARD_H - 265),
        ("Date de l'examen professionnel :", data['date_prof'], CARD_H - 295),
        ("Date de l'examen médical :", data['date_med'], CARD_H - 325),
        ("Date de l'examen psychotechnique :", data['date_psy'], CARD_H - 355),
    ]

    for label, val, y in labels_y:
        c.setFont(FONT, 11)
        c.drawString(180, y, label)
        c.setFont(FONT, 11)
        c.drawString(380, y, str(val))

    # 6. الجهة اليمنى (Engins, Sites, Manœuvre)
    # Engins
    c.setFont(FONT, 12)
    c.drawCentredString(625, CARD_H - 40, excel_info['titre_engins'])
    lines_engins = wrap_text(data['engins'], 30)
    ey = CARD_H - 75
    for l in lines_engins:
        c.drawCentredString(625, ey, l)
        ey -= 18

    # Manœuvre
    if data['manoeuvre']:
        c.setFont(FONT, 12)
        c.drawCentredString(625, 240, "Autorisé pour la Manœuvre du")
        c.drawCentredString(625, 180, data['manoeuvre'])

    # Sites
    c.setFont(FONT, 12)
    c.drawCentredString(875, CARD_H - 40, "Autorisé aux sites / lignes suivants")
    lines_sites = wrap_text(data['sites'], 22)
    sy = CARD_H - 75
    for l in lines_sites:
        c.drawCentredString(875, sy, l)
        sy -= 18

    # 7. التنبيهات الحمراء الفتية
    c.setFont(FONT, 9)
    c.setFillColorRGB(0.8, 0, 0)
    c.drawString(510, 95, "• Cette carte doit être présentée à tout contrôle;")
    c.drawString(510, 75, "• Doit être restituée en cas de retrait temporaire ou définitif des fonctions de sécurité;")
    c.drawString(510, 55, "• La fonction de sécurité à laquelle vous êtes habilité ne peut s'exercer qu'en pleine possession de vos facultés.")

    c.save()
    buffer.seek(0)
    return buffer.getvalue()

st.title("🎴 Générateur Carte Habilitation ONCF")

selected_fonction = st.selectbox(
    "Choix de la Fonction",
    list(FUNCTIONS_CONFIG.keys())
)

excel_path = FUNCTIONS_CONFIG[selected_fonction]
excel_data = load_excel_data(excel_path)

with st.form("card_form"):
    st.subheader("1. Photo de l'Agent")
    uploaded_photo = st.file_uploader("Photo d'identité (JPG / PNG)", type=["jpg", "jpeg", "png"])

    st.subheader("2. Informations Personnelles")
    col1, col2 = st.columns(2)
    with col1:
        nom = st.text_input("Nom", "")
        matricule = st.text_input("Matricule", "")
        centre = st.text_input("Centre", "CCFTC Kénitra")
    with col2:
        prenom = st.text_input("Prénom", "")
        antenne = st.text_input("Antenne", "ACFTC Kénitra")

    st.subheader("3. Dates")
    col3, col4 = st.columns(2)
    with col3:
        date_aut = st.text_input("Date d'autorisation", "")
        date_med = st.text_input("Date examen médical", "")
    with col4:
        date_prof = st.text_input("Date examen professionnel", "")
        date_psy = st.text_input("Date examen psychotechnique", "")

    st.subheader("4. Données de la fonction (Lues depuis Excel)")
    engins = st.text_area("Engins autorisés", excel_data['engins'])
    sites = st.text_area("Sites / Lignes autorisés", excel_data['sites'])
    manoeuvre = st.text_input("Autorisé pour la Manœuvre du", excel_data['manoeuvre'])

    submit = st.form_submit_button("⚡ Générer la Carte PDF")

if submit:
    photo_bytes = uploaded_photo.read() if uploaded_photo else None

    data = {
        'nom': nom, 'prenom': prenom, 'matricule': matricule,
        'fonction': selected_fonction, 'centre': centre, 'antenne': antenne,
        'date_aut': date_aut, 'date_prof': date_prof,
        'date_med': date_med, 'date_psy': date_psy,
        'engins': engins, 'manoeuvre': manoeuvre, 'sites': sites
    }

    pdf_data = generate_pdf_bytes(data, photo_bytes, excel_data)

    st.success("Carte générée avec succès ! 🎉")
    st.download_button(
        label="📥 Télécharger le PDF de la Carte",
        data=pdf_data,
        file_name=f"Carte_{nom}_{matricule}.pdf",
        mime="application/pdf"
    )
